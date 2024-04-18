import tkinter
from tkinter import messagebox

from PyQt5.QtWidgets import QFileDialog
from PyQt5 import uic
import pandas as pd
import ctypes
from datetime import datetime
from querys.dao_generic import Generic_DAO
from querys.dao import Dao
from openpyxl import load_workbook
import os


class ImportarExcel:
    gedao = Generic_DAO()
    dao = Dao()
    versionamento = 1
    files = ""

    lista_ID = dao.ids_existentes()

    def __init__(self):
        #self.caminho = Caminho.caminho_ui()
        #self.dadosUBS = uic.loadUi(self.caminho + "formulario_insercao_ubs.ui")
        #self.dadosUBS.importar_excel.clicked.connect(self.importar_ubs)
        #self.dadosUBS.comboBox_ID.addItems(self.lista_ID)
        # self.importarUBS = uic.loadUi("formulario_insercao_ubs.ui")
        # self.importarUBS.botaoPesquisar.clicked.connect(self.pesquisar)
        # self.importarUBS.botaoImportar.clicked.connect(self.importar_excel_ubs)
        # self.importarUBS.botaoCancelar.clicked.connect(self.cancelar)
        self.importar_excel_ubs()

    """
    def cancelar(self):
        self.importarUBS.close()

    def pesquisar(self):
        self.files = QFileDialog.getOpenFileName()
        print(self.files)

        self.path = self.files[0]
        print(self.path)

        self.importarUBS.endereco_excel.setText(self.path)
        print(self.importarUBS.endereco_excel.text())

    """
    def inserir_id_excel(self):
        try:
            wb = load_workbook(self.path)
            sheets = wb.sheetnames
            geral = wb[sheets[0]]
            # Then update as you want it
            geral.cell(row=60, column=2).value = 'ID: ' + self.id  # This will change the cell(2,4) to 4
            if not os.path.exists("C:/Temp/UBS"):
                os.makedirs("C:/Temp/UBS")
            self.caminho = 'C:/Temp/UBS/' + str(self.id) + '.xlsx'
            print(self.caminho)
            wb.save(self.caminho)
        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - inserir_id_excel", 0)

    """
    def atualizacao_combo_box(self):
        self.dadosUBS.comboBox_ID.clear()
        self.lista_ID = self.dao.ids_existentes()
        self.dadosUBS.comboBox_ID.addItems(self.lista_ID)

    
    def criar_verificar_id_versionamento(self):
        if str(self.dadosUBS.comboBox_ID.currentText()) == "Não existe ID":
            # criar um ID a partir do dia e hora
            self.id = str(datetime.now()).replace('-', '').replace(':', '').replace(' ', '')[0:13]
            print(self.id)
            self.versionamento = 1
        else:
            self.id = str(self.dadosUBS.comboBox_ID.currentText())
            print(self.id)
            self.versionamento = int(self.dao.consultar_versionamento(self.id)) + 1
            
    """

    def criar_id(self):
        try:
            # criar um ID a partir do dia e hora
            self.id = str(datetime.now()).replace('-', '').replace(':', '').replace(' ', '')[0:13]
            print(self.id)
            self.versionamento = 1
        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - criar_id", 0)

    def importar_excel_ubs(self):
        try:
            print("deu certo")
            #self.dadosUBS.close()
            #self.criar_verificar_id_versionamento()
            self.criar_id()
            self.enderecoExcel = ""
            self.files = QFileDialog.getOpenFileName()
            self.path = self.files[0]
            self.enderecoExcel = str(self.path)
            if self.enderecoExcel != "":
                print("Foi digitado", self.enderecoExcel)
                self.dataframeUBS = pd.read_excel(self.enderecoExcel, usecols="A:B", header=None)
                self.dataframeseries = pd.read_excel(self.enderecoExcel, sheet_name=1, header=None)

                print(self.dataframeUBS)
                print(self.dataframeseries)

                self.tratar_dataframes_ubs()
                self.ajustar_condicoes_series()
                self.inserir_dados_bd_ubs()

                #self.importarUBS.close()

                #self.atualizacao_combo_box()
            else:
                ctypes.windll.user32.MessageBoxW(0, "Você não selecionou o arquivo corretamente", "Atenção", 0)
        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - importar_excel_ubs", 0)

    def inserir_dados_bd_ubs(self):
        try:

            cnpj = str(self.listaUBS[0]['CNPJ Emissor/ Originador/ Cedente / Devedor *'])
            valor = str(self.listaUBS[0]['Valor da emissão *'])

            print(cnpj, valor)

            retorno = self.dao.consultar_dados_base(cnpj, valor)
            print(retorno)

            if retorno:
                window = tkinter.Tk()
                window.wm_withdraw()
                result = messagebox.askquestion("Atenção",
                                                "Já existem dados inseridos na base para esse CNPJ emissor com"
                                                "o mesmo valor. Tem certeza que deseja inserir novamente?",
                                                icon='question')
                if result == 'yes':
                    self.validacao_insercao_dados_gerais_ubs()
                else:
                    ctypes.windll.user32.MessageBoxW(0, "Upload cancelado pelo usuário", "Atenção", 0)

            else:
                self.validacao_insercao_dados_gerais_ubs()


        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - inserir_dados_bd_ubs", 0)

    def tratar_dataframes_ubs(self):
        try:
            self.dataframeUBS.loc[-1] = ['versionamento', self.versionamento]
            self.dataframeUBS.index = self.dataframeUBS.index + 1
            self.dataframeUBS = self.dataframeUBS.sort_index()

            print(self.dataframeUBS)

            self.dataframeUBS.loc[-1] = ['ID', self.id]
            self.dataframeUBS.index = self.dataframeUBS.index + 1
            self.dataframeUBS = self.dataframeUBS.sort_index()

            print(self.dataframeUBS)

            # transformar linhas em coluna e colunas em linhas
            self.dataframeUBS_transposed = self.dataframeUBS.T
            print(self.dataframeUBS_transposed)

            # Transformar a primeira linha em coluna e deletar a primeira linha
            self.dataframeUBS_transposed = self.dataframeUBS_transposed.drop(columns=[2, 61])
            self.dataframeUBS_transposed.columns = self.dataframeUBS_transposed.iloc[0]
            self.dataframeUBSFinal = self.dataframeUBS_transposed.iloc[1:]
            print(self.dataframeUBSFinal)

            # necessário transformar os dados em integer ou boolean antes de mandar para o bd
            self.ajustar_condicoes_gerais()
            print(self.dataframeUBSFinal)

            # transformar dataframe em lista
            self.listaUBS = self.dataframeUBSFinal.to_dict('records')
            print(self.listaUBS)
        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - tratar_dataframes_ubs", 0)

    def ajustar_condicoes_series(self):
        try:
            print(self.dataframeseries)

            self.dataframeseries.replace(to_replace="Debênture Convencional", value=7, inplace=True)
            self.dataframeseries.replace(to_replace="Debênture Infraestrutura", value=8, inplace=True)
            self.dataframeseries.replace(to_replace="Não se aplica", value=9, inplace=True)


            self.dataframeseries.replace(to_replace="Mensal", value=1, inplace=True)
            self.dataframeseries.replace(to_replace="Trimestral", value=2, inplace=True)
            self.dataframeseries.replace(to_replace="Semestral", value=3, inplace=True)
            self.dataframeseries.replace(to_replace="Anual", value=4, inplace=True)

            self.dataframeseries.replace(to_replace="%CDI", value=1, inplace=True)
            self.dataframeseries.replace(to_replace="CDI+", value=2, inplace=True)
            self.dataframeseries.replace(to_replace="Dólar", value=3, inplace=True)
            self.dataframeseries.replace(to_replace="IPCA", value=4, inplace=True)
            self.dataframeseries.replace(to_replace="NTNB", value=5, inplace=True)
            self.dataframeseries.replace(to_replace="Prefixado", value=6, inplace=True)

            self.dataframeseries.replace(to_replace="Profissional", value=1, inplace=True)
            self.dataframeseries.replace(to_replace="Qualificado", value=2, inplace=True)
            self.dataframeseries.replace(to_replace="Público Geral", value=3, inplace=True)
            self.dataframeseries.replace(to_replace="Qualificado, somente mediante registro de emissor", value=4, inplace=True)
            self.dataframeseries.replace(to_replace="Qualificado, após lock up de 3 meses", value=5, inplace=True)
            self.dataframeseries.replace(to_replace="Qualificado, após lock up de 6 meses", value=6, inplace=True)
            self.dataframeseries.replace(to_replace="Público Geral, somente mediante registro de emissor", value=7, inplace=True)
            self.dataframeseries.replace(to_replace="Público Geral, após lock up de 6 meses", value=8, inplace=True)
            self.dataframeseries.replace(to_replace="Público Geral, após lock up de 1 ano", value=9, inplace=True)

            print(self.dataframeseries)
        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - ajustar_condicoes_series", 0)

    def ajustar_condicoes_gerais(self):
        try:
            print(self.dataframeUBSFinal)
            # ATIVOS
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Ativo *'] == 'Debêntures', 'Tipo de Ativo *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Ativo *'] == 'Notas Comerciais', 'Tipo de Ativo *'] = 2
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Ativo *'] == 'CRI', 'Tipo de Ativo *'] = 3
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Ativo *'] == 'CRA', 'Tipo de Ativo *'] = 4
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Ativo *'] == 'FIDC', 'Tipo de Ativo *'] = 5
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Ativo *'] == 'Debêntures Securitizadas', 'Tipo de Ativo *'] = 6

            # Resolvência
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Revolvência'] == 'Sim', 'Revolvência'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Revolvência'] == 'Não', 'Revolvência'] = '0'

            # Subordinação
            print(self.dataframeUBSFinal.iloc[0]['Tipo de Ativo *'])
            if self.dataframeUBSFinal.iloc[0]['Tipo de Ativo *'] in [1, 2]:
                self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Subordinação'] == 'Sim', 'Subordinação'] = '1'
                self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Subordinação'] == 'Não', 'Subordinação'] = '0'

            else:
                self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Subordinação *'] == 'Sim', 'Subordinação *'] = '1'
                self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Subordinação *'] == 'Não', 'Subordinação *'] = '0'

            # Seguro
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Seguro'] == 'Sim', 'Seguro'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Seguro'] == 'Não', 'Seguro'] = '0'

            # Lastro
            if self.dataframeUBSFinal.iloc[0]['Tipo de Ativo *'] in [1, 2]:
                self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Lastro'] == 'Direitos Creditórios', 'Lastro'] = 1
                self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Lastro'] == 'Título de Dívida', 'Lastro'] = 2
            else:
                self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Lastro *'] == 'Direitos Creditórios', 'Lastro *'] = 1
                self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Lastro *'] == 'Título de Dívida', 'Lastro *'] = 2

            # Lei 12.431
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Lei 12.431 *'] == 'Sim', 'Lei 12.431 *'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Lei 12.431 *'] == 'Não', 'Lei 12.431 *'] = '0'

            # Conversiveis em ações
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Conversível em Ações *'] == 'Sim', 'Conversível em Ações *'] = '1'
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Conversível em Ações *'] == 'Não', 'Conversível em Ações *'] = '0'

            # ESG
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['ESG *'] == 'Sim', 'ESG *'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['ESG *'] == 'Não', 'ESG *'] = '0'

            # Emissor
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Emissor *'] == 'EGEM', 'Tipo de Emissor *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Emissor *'] == 'EFRF', 'Tipo de Emissor *'] = 2
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Emissor *'] == 'Emissor em fase pré-operacional', 'Tipo de Emissor *'] = 3
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Emissor *'] == 'SPAC', 'Tipo de Emissor *'] = 4
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Emissor *'] == 'SPE', 'Tipo de Emissor *'] = 5
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Emissor *'] == 'S.A. capital aberto', 'Tipo de Emissor *'] = 6
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Emissor *'] == 'S.A. capital fechado', 'Tipo de Emissor *'] = 7
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Emissor *'] == 'S.A. categoria B', 'Tipo de Emissor *'] = 8
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Tipo de Emissor *'] == 'LTDA.', 'Tipo de Emissor *'] = 9


            # Coordenador
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Nome Coordenador *'] == 'UBS BB', 'Nome Coordenador *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Nome Coordenador *'] == 'BB BI', 'Nome Coordenador *'] = 2

            # Legislação CVM
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Legislação CVM *'] == 'RCVM 160', 'Legislação CVM *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Legislação CVM *'] == 'RCVM 60', 'Legislação CVM *'] = 2
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Legislação CVM *'] == 'RCVMs 160 e 60', 'Legislação CVM *'] = 3

            # Rito de Registro
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rito de Registro *'] == 'Automático', 'Rito de Registro *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rito de Registro *'] == 'Ordinário', 'Rito de Registro *'] = 2

            # Documento Oferta
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Documentos da Oferta *'] == 'Formulário Eletrônico', 'Documentos da Oferta *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Documentos da Oferta *'] == 'Prospecto', 'Documentos da Oferta *'] = 2
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Documentos da Oferta *'] == 'Lâmina', 'Documentos da Oferta *'] = 3
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Documentos da Oferta *'] == 'Prospecto e Lâmina', 'Documentos da Oferta *'] = 4

            # Regime de Colocação
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Regime de Colocação *'] == 'Garantia Firme', 'Regime de Colocação *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Regime de Colocação *'] == 'Melhores Esforços', 'Regime de Colocação *'] = 2
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Regime de Colocação *'] == 'Garantia Firma e Melhores Esforços', 'Regime de Colocação *'] = 3

            # Lote Adicional
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Lote Adicional? *'] == 'Sim', 'Lote Adicional? *'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Lote Adicional? *'] == 'Não', 'Lote Adicional? *'] = '0'

            # Rating Emissora
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Rating da Emissora *'] == 'Não possui', 'Rating da Emissora *'] = 0
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Rating da Emissora *'] == 'Aaa/AAA/AAA', 'Rating da Emissora *'] = 1
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Rating da Emissora *'] == 'Aa1/AA+/AA+', 'Rating da Emissora *'] = 2
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Rating da Emissora *'] == 'Aa2/AA/AA', 'Rating da Emissora *'] = 3
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Rating da Emissora *'] == 'Aa3/AA-/AA-', 'Rating da Emissora *'] = 4
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating da Emissora *'] == 'A1/A+/A+', 'Rating da Emissora *'] = 5
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating da Emissora *'] == 'A2/A/A', 'Rating da Emissora *'] = 6
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating da Emissora *'] == 'A3/A-/A-', 'Rating da Emissora *'] = 7
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Rating da Emissora *'] == 'Baa1/BBB+/BBB+', 'Rating da Emissora *'] = 8
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Rating da Emissora *'] == 'Baa2/BBB/BBB', 'Rating da Emissora *'] = 9
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Rating da Emissora *'] == 'Baa3/BBB-/BBB-', 'Rating da Emissora *'] = 10

            # Rating Oferta
            self.dataframeUBSFinal.loc[
                self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'Não haverá', 'Rating Mínimo da Oferta *'] = 11
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'Aaa/AAA/AAA', 'Rating Mínimo da Oferta *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'Aa1/AA+/AA+', 'Rating Mínimo da Oferta *'] = 2
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'Aa2/AA/AA', 'Rating Mínimo da Oferta *'] = 3
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'Aa3/AA-/AA-', 'Rating Mínimo da Oferta *'] = 4
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'A1/A+/A+', 'Rating Mínimo da Oferta *'] = 5
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'A2/A/A', 'Rating Mínimo da Oferta *'] = 6
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'A3/A-/A-', 'Rating Mínimo da Oferta *'] = 7
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'Baa1/BBB+/BBB+', 'Rating Mínimo da Oferta *'] = 8
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'Baa2/BBB/BBB', 'Rating Mínimo da Oferta *'] = 9
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Rating Mínimo da Oferta *'] == 'Baa3/BBB-/BBB-', 'Rating Mínimo da Oferta *'] = 10

            # Probabilidade Colocação
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Probabilidade de colocação *'] == 'Muito Baixa', 'Probabilidade de colocação *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Probabilidade de colocação *'] == 'Baixa', 'Probabilidade de colocação *'] = 2
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Probabilidade de colocação *'] == 'Média', 'Probabilidade de colocação *'] = 3
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Probabilidade de colocação *'] == 'Alta', 'Probabilidade de colocação *'] = 4
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Probabilidade de colocação *'] == 'Muito Alta', 'Probabilidade de colocação *'] = 5

            # Haverá esforço de distribuição junto ao investidor PF no BB
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Haverá esforço de distribuição junto ao investidor PF no BB? *'] == 'Sim', 'Haverá esforço de distribuição junto ao investidor PF no BB? *'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Haverá esforço de distribuição junto ao investidor PF no BB? *'] == 'Não', 'Haverá esforço de distribuição junto ao investidor PF no BB? *'] = '0'

            # Terá garantia?
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Terá garantia? *'] == 'Sim', 'Terá garantia? *'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Terá garantia? *'] == 'Não', 'Terá garantia? *'] = '0'

            # Permite resgate antecipado?
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Permite resgate antecipado? *'] == 'Sim', 'Permite resgate antecipado? *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Permite resgate antecipado? *'] == 'Não', 'Permite resgate antecipado? *'] = 0
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Permite resgate antecipado? *'] == 'a ser definido', 'Permite resgate antecipado? *'] = 2

            # Market Flex/Mac Clause
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Market Flex e Mac Clause *'] == 'Sim', 'Market Flex e Mac Clause *'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Market Flex e Mac Clause *'] == 'Não', 'Market Flex e Mac Clause *'] = '0'

            # Fee Canal - Forma de Cálculo
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Fee Canal - Forma de Cálculo *'] == 'Flat', 'Fee Canal - Forma de Cálculo *'] = 1
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Fee Canal - Forma de Cálculo *'] == 'Ao ano', 'Fee Canal - Forma de Cálculo *'] = 2

            # Balanço Auditado
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Balanço Auditado *'] == 'Sim', 'Balanço Auditado *'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Balanço Auditado *'] == 'Não', 'Balanço Auditado *'] = '0'

            # Balanço Divulgado
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Balanço Divulgado *'] == 'Sim', 'Balanço Divulgado *'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['Balanço Divulgado *'] == 'Não', 'Balanço Divulgado *'] = '0'

            # RFP
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['RFP (Request for Proposal)? *'] == 'Sim', 'RFP (Request for Proposal)? *'] = '1'
            self.dataframeUBSFinal.loc[self.dataframeUBSFinal['RFP (Request for Proposal)? *'] == 'Não', 'RFP (Request for Proposal)? *'] = '0'

        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - ajustar_condicoes_gerais", 0)

    def gerar_lista_dados_series(self):
        try:
            #num_colunas = self.dataframeseries

            linhas = [1, 14, 27, 40, 53]
            colunas = [3, 6, 9, 12, 15]

            self.lista_series = []
            self.serie = 1

            for x in linhas:
                for y in colunas:
                    if str(self.dataframeseries.iat[x,0]) != 'nan' and str(self.dataframeseries.iat[x,y]) != 'nan':
                        self.series = {'id': self.id, 'versionamento': self.versionamento, 'serie': self.serie, 'tipo': self.dataframeseries.iat[x, 0],
                                            'prazo': self.dataframeseries.iat[x, y], 'carencia': self.dataframeseries.iat[x+1, y],
                                            'pag_principal': self.dataframeseries.iat[x+2, y], 'ini_pag_principal': self.dataframeseries.iat[x+3, y],
                                            'carencia_juros': self.dataframeseries.iat[x+4, y], 'pag_juros': self.dataframeseries.iat[x+5, y],
                                            'indexador': self.dataframeseries.iat[x+6, y], 'ref_ntnb': self.dataframeseries.iat[x+7, y],
                                            'taxa': self.dataframeseries.iat[x+8, y], 'pub_merc_primario': self.dataframeseries.iat[x+9, y],
                                            'pub_merc_secundario': self.dataframeseries.iat[x+10, y]}

                        self.serie = self.serie + 1

                        print(self.series)
                        self.lista_series.append(self.series)

            print(self.lista_series)
            return self.lista_series

        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - gerar_lista_dados_series", 0)

    def inserir_dados_bd_series(self):
        try:
            self.serie = self.serie - 1 # necessário remover 1, pois quando não há mais séries para serem incluidas, ele já somou mais 1 para o proximo
            self.c = 0
            for item in self.lista_series:
                x = self.dao.inserir_dados_series(item)
                if x == 1:
                    self.c = self.c + 1
        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - inserir_dados_bd_series", 0)

    def validacao_insercao_dados_gerais_ubs(self):
        try:
            self.b = 0
            self.c = 0
            qt_lista_series = len(self.gerar_lista_dados_series())
            print(qt_lista_series)
            print(self.listaUBS)
            if qt_lista_series == self.listaUBS[0]['Séries *']:
                self.a = self.dao.inserir_dados_gerais(self.listaUBS)
                if self.a > 0:
                    self.inserir_dados_bd_series()
                    print(self.c)

                # inserir os dados de lastro apenas nos casos de CRI, CRA e FDIC e o lastro for Título de Dívida
                x = self.listaUBS[0]['Tipo de Ativo *']
                if self.listaUBS[0]['Tipo de Ativo *'] in [1, 2]:
                    y = self.listaUBS[0]['Lastro']
                elif self.listaUBS[0]['Tipo de Ativo *'] in [3, 4, 5, 6]:
                    y = self.listaUBS[0]['Lastro *']

                print(x)
                print(y)
                if x in [3, 4, 5,
                         6] and y == 1:  # Tipo de ativo: CRI, CRA, FDIC ou Debêntures Securitizadas e Lastro: Direitos Creditórios

                    self.b = self.dao.inserir_dados_lastro(self.listaUBS)

                # Validação e mensagem para o usuário saber se os dados foram inseridos corretamente
                if self.a > 0 and self.b > 0:  # CRI, CRA, FIDC e Debentures Securitizadas
                    ctypes.windll.user32.MessageBoxW(0, f"ID: {self.id}"
                                                        f"\nDados gerais da UBS inseridos com sucesso"
                                                        f"\nDados referente aos lastros inseridos com sucesso"
                                                        f"\n Inseridos {self.c} série(s) do total de {self.serie}",
                                                     f"CRI, CRA,FIDC, Debêntures Secutizadoras", 0)
                    self.inserir_id_excel()

                    ctypes.windll.user32.MessageBoxW(0, f"Planilha com ID salva em: "
                                                        f"\n {self.caminho} ", "Atenção", 0)

                elif self.a > 0 and self.b == 0 and x in [1, 2]:
                    ctypes.windll.user32.MessageBoxW(0, f"ID: {self.id}"
                                                        f"\nDados gerais da UBS inseridos com sucesso"
                                                        f"\nInseridos {self.c} série(s) do total de {self.serie}",
                                                     "Debênture, Nota Promissoria",
                                                     0)
                    self.inserir_id_excel()

                    ctypes.windll.user32.MessageBoxW(0, f"Planilha com ID salva em: "
                                                        f"\n {self.caminho}", "Atenção", 0)

                elif self.a > 0 and self.b == 0 and x in [3, 4, 5, 6] and y == 1:
                    ctypes.windll.user32.MessageBoxW(0, f"ID: {self.id}"
                                                        f"\nDados gerais da UBS inseridos com sucesso, "
                                                        f"\n entretanto não foi possivel inserir dados do lastro. "
                                                        f"\n Verifique as informações da planilha, pois trata-se de"
                                                        f"\n CRI, CRA, FIDC ou Debêntures Secutizadoras com lastro"
                                                        f"\n em Direitos Creditórios"
                                                        f"\n Inseridos {self.c} série(s) do total de {self.serie}",
                                                     "CRI, CRA, FIDC, Debêntures Secutizadoras", 0)
                    self.inserir_id_excel()

                    ctypes.windll.user32.MessageBoxW(0, f"Planilha com ID salva em: "
                                                        f"\n {self.caminho} ", "Atenção", 0)

                elif self.a > 0 and self.b == 0 and x in [3, 4, 5, 6] and y == 2:
                    ctypes.windll.user32.MessageBoxW(0, f"ID: {self.id}"
                                                        f"\nDados gerais da UBS inseridos com sucesso, "
                                                        f"\n Não foram inseridos os dados do Lastro, "
                                                        f"\n pois trata-se de Título da Divida"
                                                        f"\n Inseridos {self.c} série(s) do total de {self.serie}",
                                                     "CRI, CRA, FIDC, Debêntures Secutizadoras", 0)
                    self.inserir_id_excel()

                    ctypes.windll.user32.MessageBoxW(0, f"Planilha com ID salva em: "
                                                        f"\n {self.caminho} ", "Atenção", 0)

                elif self.a > 0 and self.b == 0 and x in [3, 4, 5, 6] and y != y:
                    ctypes.windll.user32.MessageBoxW(0, f"ID: {self.id}"
                                                        f"\nDados gerais da UBS inseridos com sucesso, "
                                                        f"\n Não foram inseridos os dados do Lastro, "
                                                        f"\n pois os dados estão incompletos na planilha"
                                                        f"\n Inseridos {self.c} série(s) do total de {self.serie}",
                                                     "CRI, CRA, FIDC, Debêntures Secutizadoras", 0)
                    self.inserir_id_excel()

                    ctypes.windll.user32.MessageBoxW(0, f"Planilha com ID salva em: "
                                                        f"\n {self.caminho} ", "Atenção", 0)

                elif self.a == 0 and self.b == 0:
                    ctypes.windll.user32.MessageBoxW(0, "Não foi possivel fazer a inserção dos dados. "
                                                        "\n Verifique se as informações foram "
                                                        "\n preenchidas corretamente na planilha"
                                                        f"\n Não foram inseridas as series", "Atenção", 0)
            else:
                ctypes.windll.user32.MessageBoxW(0, "O número de séries informado na aba geral não "
                                                    "\n corresponde ao número existente na aba Séries. "
                                                    "\n Verifique as informações e as corrija antes de "
                                                    f"\n subir o arquivo novamente.", "Atenção", 0)

        except Exception as e:
            retorno = "Erro: {}".format(e.args)
            print(retorno)
            ctypes.windll.user32.MessageBoxW(0, retorno, "Erro - validacao_insercao_dados_gerais_ubs", 0)